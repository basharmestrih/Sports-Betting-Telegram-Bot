from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler, MessageHandler, ConversationHandler,
    ContextTypes, filters
)
from PIL import Image, ImageFilter
import pytesseract
import re
import os
import pandas as pd
import openpyxl
import random
import string
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackContext
from selenium import webdriver
from seleniumbase import Driver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from selenium.webdriver.common.keys import Keys

COLLECTING_INPUTS = range(1)
COLLECTING_DATA = range(1)


# Define  variables
user_payment_number = {}
user_deposit_amount = {}
# generate referral number
def generate_referral_code(length=8):
    letters_and_digits = string.ascii_letters + string.digits
    return ''.join(random.choice(letters_and_digits) for _ in range(length))




# State definitions for ConversationHandler
ASK_REFERRAL, GET_REFERRAL,NO_REFERRAL = range(3)


# Function to start the bot and ask about referral code
async def start(update: Update, context: CallbackContext) -> int:
    df = pd.read_excel('Book1.xlsx')
    user = update.message.from_user
    username = user.username
    if username in df['username'].values:
        return await show(update, context)
    else:
        referral_message = "هل تملك رقم احالة من احد اضدقاؤك؟"

        # Create the referral keyboard (Yes / No)
        referral_keyboard = [
            [InlineKeyboardButton("نعم", callback_data='yes_referral')],
            [InlineKeyboardButton("لا", callback_data='no_referral')]
        ]
        reply_markup = InlineKeyboardMarkup(referral_keyboard)

        # Ask user if they have a referral code
        await update.message.reply_text(referral_message, reply_markup=reply_markup)


    return ASK_REFERRAL



# Callback to handle the user's referral answer
async def ask_referral(update: Update, context: CallbackContext) -> int:
    query = update.callback_query
    await query.answer()

    if query.data == 'yes_referral':
        await query.edit_message_text("الرجاء إدخال رقم الاحالة:")
        return GET_REFERRAL  # Move to the next state to get the referral code

    elif query.data == 'no_referral':
        # Proceed without referral code, show main menu
        await no_referral(update, context)
        return NO_REFERRAL


# Function to get the referral code

async def get_referral(update: Update, context: CallbackContext) -> int:
    referral_code = update.message.text
    context.user_data['referral_code'] = referral_code  # Save referral code

    # Path to the Excel file
    excel_file = 'Book1.xlsx'

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    referral_column = 3
    username_column = 1  # Column index for usernames
    new_row = []  # List to store new row data

    # Check if the referral code exists in the column
    referral_exists = False
    for row in ws.iter_rows(min_col=referral_column, max_col=referral_column, values_only=True):
        if row[0] == referral_code:
            referral_exists = True
            break

    if referral_exists:
        await update.message.reply_text(f"تم حفظ رقم الاحالة: {referral_code}")

        # Generate a new referral code (for demonstration purposes, use UUID)
        new_referral_code = generate_referral_code()
        # Retrieve the current username of the user
        username = update.message.from_user.username

        # Add new user details to a new row in Excel
        new_row = [username, "", new_referral_code]  # Assuming the columns are: username, some column, referral code, new referral code

        ws.append(new_row)  # Append the new row to the worksheet
        wb.save(excel_file)  # Save the changes to the Excel file

        # Proceed to show the main menu
        await show(update, context)

    else:
        await update.message.reply_text("لقد ادخلت رمز دعوى خاطئ")

    return ConversationHandler.END



async def no_referral(update: Update, context: CallbackContext) -> None:
    excel_file = 'Book1.xlsx'
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    new_referral_code = generate_referral_code()
    # Retrieve the current username of the user
    if update.message and update.message.from_user:
        username = update.message.from_user.username
    elif update.callback_query and update.callback_query.from_user:
        username = update.callback_query.from_user.username
    else:
        username = "Unknown"

    # Add new user details to a new row in Excel
    new_row = [username, "",
               new_referral_code]  # Assuming the columns are: username, some column, referral code, new referral code

    ws.append(new_row)  # Append the new row to the worksheet
    wb.save(excel_file)  # Save the changes to the Excel file

    # Proceed to show the main menu
    await show(update, context)



# Function to show the main menu
async def show(update: Update, context: CallbackContext) -> None:
    welcome_message = (
        "أهلا بك في بوت ايتشانسي\n"
        "يمكنك هذا البوت من إنشاء حساب أو تعبئة رصيد والاطلاع على سجل الرهانات الخاصة بك "
    )

    keyboard = [
        [InlineKeyboardButton("🎓انشاء حساب ايتشانسي", callback_data='create_account')],
        [InlineKeyboardButton(" تعبئة  اينشانسي", callback_data='deposit')],
        [InlineKeyboardButton("سحب اينشانسي", callback_data='withdraw')],
        [InlineKeyboardButton("رصيدي", callback_data='balance')],
        [InlineKeyboardButton("كود جائزة", callback_data='reward')],
        [InlineKeyboardButton("جروب الدعم", url="https://t.me/+JkZ3-g6U7oM0NGQ0")],
        [InlineKeyboardButton("📊 اايداع الرصيد في ايتشانسي", callback_data='ich_deposit')],
        [InlineKeyboardButton("⚙️ السحل", callback_data='history')],
        [InlineKeyboardButton("💎 الدليل الشامل", callback_data='subscription')],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    # Check if the update comes from a message or a callback query
    if update.message:
        await update.message.reply_text(welcome_message, reply_markup=reply_markup)
    elif update.callback_query:
        await update.callback_query.message.edit_text(welcome_message, reply_markup=reply_markup)


async def create_account(update: Update, context: CallbackContext) -> int:
    query = update.callback_query
    await query.answer()
    await query.edit_message_text("ادخل اسم المستخدم، كلمة المرور،الايميل, كل منهما في رسالة منفصلة:")
    await collect_inputs(update, context)


    return COLLECTING_INPUTS
# Collect inputs function
async def collect_inputs(update: Update, context: CallbackContext) -> int:
    # Initialize the inputs list if it doesn't exist yet
    if 'inputs' not in context.user_data:
        context.user_data['inputs'] = []

    # Append the user input to the list
    context.user_data['inputs'].append(update.message.text)

    # Once we have collected all three inputs (username, password, email)
    if len(context.user_data['inputs']) == 3:
        username, password, email = context.user_data['inputs']

        # Get the Telegram username of the user
        telegram_username = update.message.from_user.username

        # Reply to the user
        await update.message.reply_text("حاري انشاء الحساب الخاص بك يرجى الانتظار!")

        # Update the Excel file with the collected inputs
        wb = openpyxl.load_workbook('Book1.xlsx')
        sheet = wb.active

        # Iterate over the rows to find the row with the matching Telegram username
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
            cell = row[0]  # The first column contains Telegram usernames
            if cell.value == telegram_username:
                # If found, update the 4th and 5th columns (D and E) in the same row
                sheet.cell(row=cell.row, column=4).value = username  # 4th column for username
                sheet.cell(row=cell.row, column=5).value = password  # 5th column for password
                break

        # Save the workbook
        wb.save('Book1.xlsx')

        # For debugging purposes

        # Clear the inputs after processing
        context.user_data['inputs'].clear()

        page = ChromiumPage()

        try:
            # Open the target website
            page.get('https://agents.ichancy.com/')

            time.sleep(1)
            page.get('https://agents.ichancy.com/players/players')
            time.sleep(1)
            try:
                username_field = page.ele(
                    'xpath://*[@id="root"]/div/div[3]/div[2]/div[2]/div[2]/form/div[1]/div/label/div[1]/input',
                    timeout=5
                )
                password_field = page.ele(
                    'xpath://*[@id="root"]/div/div[3]/div[2]/div[2]/div[2]/form/div[2]/div/label/div[1]/input',
                    timeout=5
                )

                # If username and password fields are located, fill them
                if username_field and password_field:
                    username_field.input('thelegend@agent.nsp')
                    password_field.input('Aa990@990\n')  # This will submit the form after entering the password
                    sleep(5)
            except Exception:
                # If username/password field is not found, skip filling them
                print("Username or password field not found, skipping login step.")
                pass

            # Add player button
            element = page.ele('css:.btn.playersActionButton-bc')
            element.click()

            time.sleep(1)

            # Fill in player information
            name = page.ele(
                'xpath:.//*[@id="root"]/div/div[4]/div/div/div[2]/form/div/div/div[1]/div/div/label/div[1]/input')
            name.input('ahmed')

            mid_name = page.ele(
                'xpath:.//*[@id="root"]/div/div[4]/div/div/div[2]/form/div/div/div[2]/div/div/label/div[1]/input')
            mid_name.input('ali')

            last_name = page.ele(
                'xpath:.//*[@id="root"]/div/div[4]/div/div/div[2]/form/div/div/div[3]/div/div/label/div[1]/input')
            last_name.input('fateh')

            user_name = page.ele(
                'xpath:.//*[@id="root"]/div/div[4]/div/div/div[2]/form/div/div/div[4]/div/div/label/div[1]/input')
            user_name.input(username)

            phone = page.ele(
                'xpath:.//*[@id="root"]/div/div[4]/div/div/div[2]/form/div/div/div[5]/div/div/label/div[1]/input')
            phone.input('123456789')

            mail = page.ele(
                'xpath:.//*[@id="root"]/div/div[4]/div/div/div[2]/form/div/div/div[6]/div/div/label/div[1]/input')
            mail.input(email)

            keyword = page.ele(
                'xpath:.//*[@id="root"]/div/div[4]/div/div/div[2]/form/div/div/div[7]/div/div/label/div[1]/input')
            keyword.input(password)
            time.sleep(5)
            # Handle the dropdown selection
            dropdown_list = page.ele('xpath:.//input[@placeholder="Countries"]')
            dropdown_list.click()
            andorra_element = page.ele('xpath:.//p[@title="Andorra"]')

            andorra_element.click()

            parent_list = page.ele(
                'xpath:.//*[@id="root"]/div/div[4]/div/div/div[2]/form/div/div/div[10]/div/div[1]/div[1]/label/div[1]/input')
            parent_list.click()

            element = page.ele(
                'xpath:.//*[@id="root"]/div/div[4]/div/div/div[2]/form/div/div/div[10]/div/div[2]/div/button')
            element.click()

            # Submit registration
            reg = page.ele(
                'xpath:.//*[@id="root"]/div/div[4]/div/div/div[3]/button[2]')
            reg.click()

            time.sleep(10)

            await update.message.reply_text(f"بيانات الحساب الخاص بك\n"
                  f"اسم المستخدم: {username}\n"
                  f"كلمة السر: {password}")

        finally:
            # Close the browser
            page.quit()

        # End the conversation
        return ConversationHandler.END



#deposit
def extract_info_from_image(image_path):
    # Load the image
    image = Image.open(image_path)

    # Convert the image to grayscale for better OCR results
    gray_image = image.convert('L')

    # Enhance the image by applying a sharpening filter to make the text clearer
    enhanced_image = gray_image.filter(ImageFilter.SHARPEN)

    # Define the region to crop the first box (for the number on the left)
    first_box_region = enhanced_image.crop((0, 350, image.width // 2, 450))
    # Perform OCR on the entire image to extract all text

    # Extract all occurrences of the text after "Amount:"
    # Define the region to crop the second row (where "To:" is located)
    to_number_region = enhanced_image.crop((0, 450, image.width, 550))

    # Perform OCR on the cropped regions
    ocr_first_box = pytesseract.image_to_string(first_box_region)
    ocr_to_number = pytesseract.image_to_string(to_number_region)
    ocr_result = pytesseract.image_to_string(enhanced_image)


    # Process the OCR result to extract the desired information
    lines = ocr_first_box.splitlines()
    number = lines[0].strip() if lines else "Number not found"
    amounts = re.findall(r'Amount:\s*(\d+)', ocr_result)


    match = re.search(r'To:\s*(\d+)', ocr_to_number)
    to_number = match.group(1) if match else "Number not found"

    return number, to_number , amounts
async def deposit_withdraw(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()

    if query.data == 'deposit':
        keyboard = [
            [InlineKeyboardButton("syriatel cash", callback_data='deposit_cash')],
            [InlineKeyboardButton("payeer", callback_data='deposit_payeer')],
            [InlineKeyboardButton("Bemo bank", callback_data='deposit_bemo')],
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text("اختر طرق الدفع", reply_markup=reply_markup)
async def handle_deposit_method_selection(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()

    if query.data == 'deposit_cash':
        context.user_data['deposit_method'] = query.data.split('_')[1]

        await query.edit_message_text("قم بالتحويل إلى هذا الرقم 22443355 وادخل رقم عملية التحويل")

        context.user_data['awaiting_payment_number'] = True
    elif query.data == 'deposit_payeer':
        await query.edit_message_text("قم بالتحويل إلى هذه المحفظة P1034210265 وادخل المبلغ المودع")
        context.user_data['awaiting_payeerpayment_number'] = True
    elif query.data == 'deposit_bemo':
        await query.edit_message_text("قم بالتحويل إلى هذه المحفظة 22445566 وادخل المبلغ المودع")
        context.user_data['awaiting_bemopayment_number'] = True
#cash handle
processed_payment_numbers = []

# Cash handle function for processing payment numbers
async def message_handler2(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.message.from_user.id

    if context.user_data.get('awaiting_payment_number'):
        payment_number = update.message.text

        # Check if the payment number has already been used
        if payment_number in processed_payment_numbers:
            await update.message.reply_text("تم الايداع مسبقا باستخدام رقم التحويل المدخل")
        else:
            # Store the payment number for this user
            user_payment_number[user_id] = payment_number
            await update.message.reply_text("ادخل رقم المبلغ المراد ايداعه")
            context.user_data['awaiting_payment_number'] = False
            context.user_data['awaiting_deposit_amount'] = True

    elif context.user_data.get('awaiting_deposit_amount'):
        deposit_amount = int(update.message.text)
        if deposit_amount < 15000:
            await update.message.reply_text("المبلغ أقل من الحد الأدنى المسموح به")
        else:
            context.user_data['deposit_amount'] = deposit_amount
            await update.message.reply_text("ارسل لقطة شاشة لعملية التحويل من سيريتل كاش")
            context.user_data['awaiting_deposit_amount'] = False
            context.user_data['awaiting_screenshot'] = True


async def photo_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.message.from_user.id

    if context.user_data.get('awaiting_screenshot'):
        # Check if the message contains a photo
        if not update.message.photo:
            await update.message.reply_text("No photo found in the message. Please send a valid screenshot.")
            return

        # Download the photo
        photo_file = await update.message.photo[-1].get_file()

        # Ensure the 'downloads' directory exists
        downloads_dir = 'downloads'
        if not os.path.exists(downloads_dir):
            os.makedirs(downloads_dir)

        file_path = os.path.join(downloads_dir, f'{user_id}_screenshot.jpg')
        await photo_file.download_to_drive(file_path)

        # Process the image to extract information
        number, to_number, amounts = extract_info_from_image(file_path)

        # Extracting the values from the dictionaries
        deposit_amount = user_deposit_amount[list(user_deposit_amount.keys())[0]]
        payment_number = user_payment_number[list(user_payment_number.keys())[0]]

        # Checking the conditions
        if amounts[0] == deposit_amount and number == payment_number:
            # Update Excel file with the deposit amount
            excel_file = 'Book1.xlsx'
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            # Variables for updating balances
            user_found = False
            referrer_username = None

            # Iterate over rows to find the username and referrer
            for row_index in range(1, ws.max_row + 1):  # Start from row 2
                cell_username = ws.cell(row=row_index, column=1).value  # Username column
                referral_code = ws.cell(row=row_index, column=3).value  # Referral code column

                if cell_username == update.message.from_user.username:
                    # Update the balance for the depositing user
                    ws.cell(row=row_index, column=2, value=amounts[0])  # Balance column
                    user_found = True

                    # Find the referrer
                    if referral_code:
                        for ref_row_index in range(1, ws.max_row + 1):
                            if ws.cell(row=ref_row_index, column=3).value == referral_code:  # Referral code column
                                referrer_username = ws.cell(row=ref_row_index, column=1).value
                                break
                    break

            if user_found and referrer_username:
                # Calculate 10% of the deposit amount
                referral_reward = deposit_amount * 0.10

                # Update the referrer's balance
                for row_index in range(1, ws.max_row + 1):
                    if ws.cell(row=row_index, column=1).value == referrer_username:
                        current_balance = ws.cell(row=row_index, column=2).value or 0
                        ws.cell(row=row_index, column=2, value=current_balance + referral_reward)
                        break

            # Save changes to the Excel file
            wb.save(excel_file)

            # Notify user of successful deposit
            await update.message.reply_text("تم اضافة الرصيد بنجاح")
        else:
            await update.message.reply_text("معلومات الإيداع غير صحيحة. يرجى التحقق من التفاصيل وإعادة الإرسال.")

        # Reset the flag
        context.user_data['awaiting_screenshot'] = False

deposit_amount2 = 0

#payeer handle
async def payeer_handle(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    global deposit_amount2
    deposit_amount2 = int(update.message.text)
    if deposit_amount2 < 15000:
        await update.message.reply_text("المبلغ أقل من الحد الأدنى المسموح به")
    else:
        context.user_data['awaiting_payeerpayment_number'] = deposit_amount2
        await update.message.reply_text("ارسل لقطة شاشة لعملية التحويل من بايير")
        context.user_data['awaiting_payeerpayment_number'] = False
        context.user_data['awaiting_photo'] = True
async def payeer_photo_handle(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    username = update.message.from_user.username

    if context.user_data.get('awaiting_photo'):
        print('fhfhfh')
        # Check if the message contains a photo
        if not update.message.photo:
            await update.message.reply_text("No photo found in the message. Please send a valid screenshot.")
            return

        # Download the photo
        photo_file = await update.message.photo[-1].get_file()

        # Ensure the 'downloads' directory exists
        downloads_dir = 'downloads'
        if not os.path.exists(downloads_dir):
            os.makedirs(downloads_dir)
        file_path = os.path.join(downloads_dir, f'{username}_screenshot.jpg')
        await photo_file.download_to_drive(file_path)
        message_text = f"Amount: {deposit_amount2} SYP\nUsername: @{username}\nPayment method: Payeer"
        recipient_user_id = '5666304947'
        with open(file_path, 'rb') as photo:
            await context.bot.send_photo(chat_id=recipient_user_id, photo=photo, caption=message_text)
        await update.message.reply_text("تم اضافة طلب الايداع سيتم معالجة الطلب")
#bemo bank
deposit_amount3 = 0
async def bemo_handle(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    global deposit_amount3
    deposit_amount3 = int(update.message.text)
    if deposit_amount3 < 15000:
        await update.message.reply_text("المبلغ أقل من الحد الأدنى المسموح به")
    else:
        context.user_data['awaiting_bemopayment_number'] = deposit_amount3
        await update.message.reply_text("ارسل لقطة شاشة لعملية التحويل من تطبيق بنك بيمو")
        context.user_data['awaiting_bemopayment_number'] = False
        context.user_data['awaiting_bemophoto'] = True
async def bemo_photo_handle(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    username = update.message.from_user.username

    if context.user_data.get('awaiting_bemophoto'):
        # Check if the message contains a photo
        if not update.message.photo:
            await update.message.reply_text("No photo found in the message. Please send a valid screenshot.")
            return

        # Download the photo
        photo_file = await update.message.photo[-1].get_file()

        # Ensure the 'downloads' directory exists
        downloads_dir = 'downloads'
        if not os.path.exists(downloads_dir):
            os.makedirs(downloads_dir)
        file_path = os.path.join(downloads_dir, f'{username}_screenshot.jpg')
        await photo_file.download_to_drive(file_path)
        message_text = f"Amount: {deposit_amount3} SYP\nUsername: @{username}\nPayment method: bemo bank"
        recipient_user_id = '5666304947'
        with open(file_path, 'rb') as photo:
            await context.bot.send_photo(chat_id=recipient_user_id, photo=photo, caption=message_text)
        await update.message.reply_text("تم اضافة طلب الايداع سيتم معالجة الطلب")






#wthdraw
# withdraw
async def withdraw(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()

    if query.data == 'withdraw':
        # Show withdrawal options
        keyboard = [
            [InlineKeyboardButton("syriatel cash", callback_data='withdraw_syriatel_cash')],
            [InlineKeyboardButton("payeer", callback_data='withdraw_payeer')],
            [InlineKeyboardButton("Bemo bank", callback_data='withdraw_bemo_bank')],
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text("من فضلك اختر طريقة السحب", reply_markup=reply_markup)
async def handle_withdraw_method_selection(update: Update, context: CallbackContext) -> int:
    query = update.callback_query
    await query.answer()

    if query.data.startswith('withdraw_'):
        context.user_data['withdraw_method'] = query.data.split('_')[1]
        await query.edit_message_text("ادخل رقم التحويل الخاص بك")
        context.user_data['awaiting_wallet_number'] = True
async def collect_wallet_number(update: Update, context: CallbackContext) -> int:
    user_wallet_number = update.message.text
    username = update.effective_user.username

    if context.user_data.get('awaiting_wallet_number'):
        context.user_data['wallet_number'] = update.message.text
        await update.message.reply_text("الان قم بإدخال المبلغ المراد سحبه")
        context.user_data['awaiting_wallet_number'] = False
        context.user_data['awaiting_withdraw_amount'] = True

    # Capture and handle the withdrawal amount
    elif context.user_data.get('awaiting_withdraw_amount'):
        context.user_data['withdraw_amount'] = update.message.text
        withdraw_method = context.user_data.get('withdraw_method')
        wallet_number = context.user_data.get('wallet_number')
        withdraw_amount = float(context.user_data.get('withdraw_amount'))
        excel_data = pd.read_excel("Book1.xlsx")
        user_row = excel_data[excel_data['username'] == username]
        balance = user_row['balance'].values[0]
        if withdraw_amount < 100000:
            await update.message.reply_text("لا يمكن سحب اي مبلغ تحت 100 الف ليرة سورية")
        elif withdraw_amount > balance:
            await update.message.reply_text("هذرا رصيدك الحالي غير كافي لاتمام عملية السحب")
        else:
            new_balance = balance - withdraw_amount
            # Update the balance in the DataFrame
            excel_data.loc[excel_data['username'] == username, 'balance'] = new_balance
            # Write the updated DataFrame back to the Excel file
            excel_data.to_excel("Book1.xlsx", index=False)
            withdraw_after = withdraw_amount * 0.1
            w2 = withdraw_amount - withdraw_after
            await update.message.reply_text(f"جاري معالجة طلب السحب سيصلك مبلغ{w2} ")
            await update.message.reply_text(f'رصيدك الحالي هو {new_balance}')
            await update.message.reply_text(
                f"طريقة السحب: {withdraw_method}\n"
                f"رقم المحفظة: {wallet_number}\n"
                f"المبلغ المراد سحبه: {w2}\n"
                f"اسم المستخدم: @{username}"
            )

            # Reset the state after processing the withdrawal

        context.user_data['awaiting_withdraw_amount'] = False



#my balance
async def my_balance(update: Update, context: CallbackContext) -> None:
    query = update.callback_query
    await query.answer()  # Acknowledge the callback

    user_id = update.effective_user.username

    # Load the Excel file
    df = pd.read_excel('Book1.xlsx')

    # Check if the column name is correct (replace 'User ID' with the correct column name)
    user_row = df[df['username'] == user_id]  # Use the correct column name

    if not user_row.empty:
        balance = user_row.iloc[0]['balance']  # Ensure 'Balance' is also the correct column name
        await query.edit_message_text(f"رصيدك الحالي هو:{balance}")
    else:
        await query.edit_message_text("User not found in the record.")



# reward
async def reward(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()
    await query.edit_message_text("اهلا بك في قسم الجوائز ادخل كود الجائزة")
    context.user_data['awaiting_reward_code'] = True
valid_codes = [
    '482913', '651472', '398205', '720154', '893721',
    '125038', '670492', '385716', '142803', '579128',
    '408392', '923716', '307159', '649287', '593028',
    '184930', '572839', '614023', '853094', '392710'
]

# Function to handle the 'reward' query
async def reward_query_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()
    await query.message.reply_text("اهلا بك في قسم الجوائز ادخل كود الجائزة")
    context.user_data['awaiting_reward_code'] = True
# Function to handle the user's input (reward code)
async def handle_reward_code(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if context.user_data.get('awaiting_reward_code'):
        reward_code = update.message.text.strip()

        # Debugging: Print the entered reward code and the list of valid codes
        print(f"Entered code: {reward_code}")
        print(f"Valid codes: {valid_codes}")

        # Check if the code is valid
        if reward_code in valid_codes:
            # Load the Excel file
            df = pd.read_excel('Book1.xlsx')
            user = update.message.from_user.username

            # Check if the user exists in the Excel file
            if user in df['username'].values:
                index = df[df['username'] == user].index[0]

                # Update the balance
                df.at[index, 'balance'] += 15000

                # Save the updated Excel file
                df.to_excel('Book1.xlsx', index=False)

                # Remove the used code from the list
                valid_codes.remove(reward_code)

                # Send confirmation message
                await update.message.reply_text("مبروك لقد تم تحديث رصيدك")
            else:
                await update.message.reply_text("عذرا، لم يتم العثور على المستخدم في قاعدة البيانات")
        else:
            await update.message.reply_text("الكود غير صحيح")

        # Reset the state
        context.user_data['awaiting_reward_code'] = False





from DrissionPage import ChromiumPage
from time import sleep
def process_deposit(current_username: str, input_amount: float):
    # Initialize DrissionPage
    page = ChromiumPage()
    page.get('https://agents.ichancy.com/')
    sleep(5)

    try:
        username_field = page.ele(
            'xpath://*[@id="root"]/div/div[3]/div[2]/div[2]/div[2]/form/div[1]/div/label/div[1]/input',
            timeout=5
        )
        password_field = page.ele(
            'xpath://*[@id="root"]/div/div[3]/div[2]/div[2]/div[2]/form/div[2]/div/label/div[1]/input',
            timeout=5
        )

        # If username and password fields are located, fill them
        if username_field and password_field:
            username_field.input('thelegend@agent.nsp')
            password_field.input('Aa990@990\n')  # This will submit the form after entering the password
            sleep(5)
    except Exception:
        # If username/password field is not found, skip filling them
        print("Username or password field not found, skipping login step.")
        pass

    # Navigate to the transfer page
    transfer_button = page.ele('css:.headerNavIcon-bc.cursor-pointer .bc-icon-transfer-bold')
    transfer_button.click()
    sleep(5)

    # Locate the user search bar
    user_search_bar = page.ele(
        'xpath://*[@id="root"]/div/div[4]/div/div/div[2]/form/div[1]/div/div[1]/div/div[1]/div[1]/label/div[1]/input'
    )
    sleep(5)

    # Search for the user by username
    user_search_bar.input(current_username)
    sleep(15)

    # Select the correct user from the search results
    buttons = page.eles(
        'css:button.btn.listItem.a-minimal.s-default.f-full-width.c-default.id-start.cr-round'
    )
    for button in buttons:
        if current_username in button.text:
            button.click()
            break

    # Enter the transfer amount
    transfer_amount = page.ele(
        'xpath://*[@id="root"]/div/div[4]/div/div/div[2]/form/div[1]/div/div[2]/div/div/label/div[1]/input'
    )
    transfer_amount.input(str(input_amount))
    sleep(5)

    # Confirm the transfer
    done_button = page.ele('xpath://*[@id="root"]/div/div[4]/div/div/div[3]/button[2]')
    done_button.click()
    sleep(10)

    # Save screenshot
    sleep(3)

    # Close the browser
    page.close()

async def ich_deposit(update: Update, context: CallbackContext) -> int:
    # Ensure we have a callback query
    if update.callback_query:
        # Extract chat_id and user information from callback_query
        chat_id = update.callback_query.message.chat_id
        user = update.callback_query.from_user
        username = user.username

        # Notify user that processing has started
        await context.bot.send_message(chat_id, "We are processing your request. Please wait...")

        # Define the path to your Excel file
        excel_file_path = 'Book1.xlsx'

        # Load the Excel file into a DataFrame
        import pandas as pd
        df = pd.read_excel(excel_file_path)

        # Check if username exists in the first column
        if username in df.iloc[:, 0].values:
            # Locate the row with the username
            row = df[df.iloc[:, 0] == username]

            # Get values from column 4 and column 2
            name = row.iloc[0, 3]  # Column 4
            amount = row.iloc[0, 1]  # Column 2
            row_index = df[df.iloc[:, 0] == username].index[0]
            df.at[row_index, df.columns[1]] = 0  # Column 2
            df.to_excel(excel_file_path, index=False)
            print('ff')

            # Call the process_deposit function with the extracted values
            process_deposit(name, amount)

            # Notify user that the process is complete
            await context.bot.send_message(chat_id, "Your order is done. Thank you!")
        else:
            # Notify user if the username is not found
            await context.bot.send_message(chat_id, f"Username {username} not found in the Excel file.")

    else:
        # Handle the case where there is no callback query
        await context.bot.send_message( "This update does not contain a callback query.")

    return int



EXCEL_FILE = 'Book1.xlsx'
# Function to update the Excel file with the new balance
def update_balance_in_excel(username, amount):
    # Load the workbook and select the active sheet
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active

    # Find the row where the username matches, and update the balance
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
        current_username = row[0].value
        balance_cell = row[1]  # Assuming the balance is in the second column

        if current_username == username:
            current_balance = balance_cell.value
            new_balance = current_balance + int(amount)
            balance_cell.value = new_balance
            wb.save(EXCEL_FILE)  # Save changes to the file
            return True

    return False  # Return False if the username was not found

# Asynchronous function to handle balance updates
async def handle_balance(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if context.user_data.get('awaiting_balance', False):
        message = update.message.text
        pattern = r"Amount:\s*(\d+)\s*SYP\s*Username:\s*@(\w+)\s*Payment method:\s*(.*)"
        match = re.search(pattern, message)

        if match:
            amount = match.group(1)
            username = match.group(2)
            # Update the balance in the Excel file
            success = update_balance_in_excel(username, amount)

            if success:
                await update.message.reply_text("تمت اضافة الرصيد بنجاح")
            else:
                await update.message.reply_text(f"Username @{username} not found in the system.")

        else:
            await update.message.reply_text("Please provide the message in the correct format.")

        # Reset the flag after processing the message
        context.user_data['awaiting_balance'] = False
    else:
        # If the bot is not expecting a message, it ignores or handles messages normally
        await update.message.reply_text("This message does not follow the expected format.")

# Asynchronous function to handle the /addbalance command
async def add_balance(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    context.user_data['awaiting_balance'] = True
    await update.message.reply_text("اضف نص رسالة طلب الايداع فقط سيتم اضافة الرصيد بشكل اوتوماتيكي")

# Function to handle user messages after /addbalance is issued




# Function to handle the user's input (reward code)
#api here
def main() -> None:
    application = Application.builder().token("7016588209:AAHlKI3foDHlT07OV_XDg6XdWDjtX7XenWg").build()
    # Set up the conversation handler
    conv_handler = ConversationHandler(
        entry_points=[CallbackQueryHandler(create_account, pattern='create_account')],
        states={
            COLLECTING_INPUTS: [MessageHandler(filters.TEXT & ~filters.COMMAND, collect_inputs)],
        },
        fallbacks=[],
        per_message=False,
    )

    conv_handler2 = ConversationHandler(
        entry_points=[CommandHandler('start', start)],
        states={
            ASK_REFERRAL: [CallbackQueryHandler(ask_referral)],
            GET_REFERRAL: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_referral)],
            NO_REFERRAL: [MessageHandler(filters.TEXT & ~filters.COMMAND, no_referral)],
        },
        fallbacks=[CommandHandler('start', start)],
    )



    # Add handlers

    application.add_handler(CallbackQueryHandler(deposit_withdraw, pattern='^deposit$'))
    application.add_handler(CallbackQueryHandler(handle_deposit_method_selection, pattern='^deposit_'))
    application.add_handler(CallbackQueryHandler(withdraw, pattern='^withdraw$'))
    application.add_handler(CallbackQueryHandler(handle_withdraw_method_selection, pattern='^withdraw_'))
    application.add_handler(CallbackQueryHandler(my_balance, pattern='^balance$'))
    application.add_handler(CallbackQueryHandler(reward, pattern='^reward'))
    application.add_handler(CallbackQueryHandler(ich_deposit, pattern='^ich_deposit'))
    application.add_handler(CommandHandler("addbalance", add_balance))


    # Add both message handlers in a single line with 'elif' for collecting inputs
    application.add_handler(
        MessageHandler(
            filters.TEXT & ~filters.COMMAND,
            lambda update, context: (
                message_handler2(update, context) if 'awaiting_payment_number' in context.user_data else
                collect_wallet_number(update, context) if 'awaiting_wallet_number' in context.user_data else
                payeer_handle(update, context) if 'awaiting_payeerpayment_number' in context.user_data else
                bemo_handle(update, context) if 'awaiting_bemopayment_number' in context.user_data else
                handle_reward_code(update, context) if 'awaiting_reward_code' in context.user_data else
                collect_inputs(update, context) if 'inputs' in context.user_data else
                get_referral(update, context) if 'referral_code' in context.user_data else
                handle_balance(update, context) if 'awaiting_balance' in context.user_data else
                no_referral(update, context)
            )
        )
    )

    application.add_handler(
        MessageHandler(
            filters.PHOTO & ~filters.COMMAND,
            lambda update, context: (
                photo_handler(update, context) if 'awaiting_screenshot' in context.user_data
                else bemo_photo_handle(update, context) if 'awaiting_bemophoto' in context.user_data
                else payeer_photo_handle(update, context)

            )
        )
    )






    application.add_handler(conv_handler)
    application.add_handler(conv_handler2)

    # Start polling
    application.run_polling()

if __name__ == '__main__':
    main()












from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler, MessageHandler, ConversationHandler,
    ContextTypes, filters
)
from PIL import Image, ImageFilter
import pytesseract
import re
import os
import pandas as pd
import openpyxl
import random
import string
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackContext
from selenium import webdriver
from seleniumbase import Driver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from selenium.webdriver.common.keys import Keys

COLLECTING_INPUTS = range(1)
COLLECTING_DATA = range(1)


# Define  variables
user_payment_number = {}
user_deposit_amount = {}
# generate referral number
def generate_referral_code(length=8):
    letters_and_digits = string.ascii_letters + string.digits
    return ''.join(random.choice(letters_and_digits) for _ in range(length))




# State definitions for ConversationHandler
ASK_REFERRAL, GET_REFERRAL,NO_REFERRAL = range(3)


# Function to start the bot and ask about referral code
async def start(update: Update, context: CallbackContext) -> int:
    df = pd.read_excel('Book1.xlsx')
    user = update.message.from_user
    username = user.username
    if username in df['username'].values:
        return await show(update, context)
    else:
        referral_message = "هل تملك رقم احالة من احد اضدقاؤك؟"

        # Create the referral keyboard (Yes / No)
        referral_keyboard = [
            [InlineKeyboardButton("نعم", callback_data='yes_referral')],
            [InlineKeyboardButton("لا", callback_data='no_referral')]
        ]
        reply_markup = InlineKeyboardMarkup(referral_keyboard)

        # Ask user if they have a referral code
        await update.message.reply_text(referral_message, reply_markup=reply_markup)


    return ASK_REFERRAL



# Callback to handle the user's referral answer
async def ask_referral(update: Update, context: CallbackContext) -> int:
    query = update.callback_query
    await query.answer()

    if query.data == 'yes_referral':
        await query.edit_message_text("الرجاء إدخال رقم الاحالة:")
        return GET_REFERRAL  # Move to the next state to get the referral code

    elif query.data == 'no_referral':
        # Proceed without referral code, show main menu
        await no_referral(update, context)
        return NO_REFERRAL


# Function to get the referral code

async def get_referral(update: Update, context: CallbackContext) -> int:
    referral_code = update.message.text
    context.user_data['referral_code'] = referral_code  # Save referral code

    # Path to the Excel file
    excel_file = 'Book1.xlsx'

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    referral_column = 3
    username_column = 1  # Column index for usernames
    new_row = []  # List to store new row data

    # Check if the referral code exists in the column
    referral_exists = False
    for row in ws.iter_rows(min_col=referral_column, max_col=referral_column, values_only=True):
        if row[0] == referral_code:
            referral_exists = True
            break

    if referral_exists:
        await update.message.reply_text(f"تم حفظ رقم الاحالة: {referral_code}")

        # Generate a new referral code (for demonstration purposes, use UUID)
        new_referral_code = generate_referral_code()
        # Retrieve the current username of the user
        username = update.message.from_user.username

        # Add new user details to a new row in Excel
        new_row = [username, "", new_referral_code]  # Assuming the columns are: username, some column, referral code, new referral code

        ws.append(new_row)  # Append the new row to the worksheet
        wb.save(excel_file)  # Save the changes to the Excel file

        # Proceed to show the main menu
        await show(update, context)

    else:
        await update.message.reply_text("لقد ادخلت رمز دعوى خاطئ")

    return ConversationHandler.END



async def no_referral(update: Update, context: CallbackContext) -> None:
    excel_file = 'Book1.xlsx'
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    new_referral_code = generate_referral_code()
    # Retrieve the current username of the user
    if update.message and update.message.from_user:
        username = update.message.from_user.username
    elif update.callback_query and update.callback_query.from_user:
        username = update.callback_query.from_user.username
    else:
        username = "Unknown"

    # Add new user details to a new row in Excel
    new_row = [username, "",
               new_referral_code]  # Assuming the columns are: username, some column, referral code, new referral code

    ws.append(new_row)  # Append the new row to the worksheet
    wb.save(excel_file)  # Save the changes to the Excel file

    # Proceed to show the main menu
    await show(update, context)



# Function to show the main menu
async def show(update: Update, context: CallbackContext) -> None:
    welcome_message = (
        "أهلا بك في بوت ايتشانسي\n"
        "يمكنك هذا البوت من إنشاء حساب أو تعبئة رصيد والاطلاع على سجل الرهانات الخاصة بك "
    )

    keyboard = [
        [InlineKeyboardButton("🎓انشاء حساب ايتشانسي", callback_data='create_account')],
        [InlineKeyboardButton(" تعبئة وسحب اينشانسي", callback_data='deposit')],
        [InlineKeyboardButton("سحب اينشانسي", callback_data='withdraw')],
        [InlineKeyboardButton("رصيدي", callback_data='balance')],
        [InlineKeyboardButton("كود جائزة", callback_data='reward')],
        [InlineKeyboardButton("جروب الدعم", url="https://t.me/+JkZ3-g6U7oM0NGQ0")],
        [InlineKeyboardButton("📊 اايداع الرصيد في ايتشانسي", callback_data='ich_deposit')],
        [InlineKeyboardButton("⚙️ السحل", callback_data='history')],
        [InlineKeyboardButton("💎 الدليل الشامل", callback_data='subscription')],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    # Check if the update comes from a message or a callback query
    if update.message:
        await update.message.reply_text(welcome_message, reply_markup=reply_markup)
    elif update.callback_query:
        await update.callback_query.message.edit_text(welcome_message, reply_markup=reply_markup)


async def create_account(update: Update, context: CallbackContext) -> int:
    query = update.callback_query
    await query.answer()
    await query.edit_message_text("ادخل اسم المستخدم، كلمة المرور،الايميل, كل منهما في رسالة منفصلة:")
    await collect_inputs(update, context)


    return COLLECTING_INPUTS
# Collect inputs function
async def collect_inputs(update: Update, context: CallbackContext) -> int:
    # Initialize the inputs list if it doesn't exist yet
    if 'inputs' not in context.user_data:
        context.user_data['inputs'] = []

    # Append the user input to the list
    context.user_data['inputs'].append(update.message.text)

    # Once we have collected all three inputs (username, password, email)
    if len(context.user_data['inputs']) == 3:
        username, password, email = context.user_data['inputs']

        # Get the Telegram username of the user
        telegram_username = update.message.from_user.username

        # Reply to the user
        await update.message.reply_text("حاري انشاء الحساب الخاص بك يرجى الانتظار!")

        # Update the Excel file with the collected inputs
        wb = openpyxl.load_workbook('Book1.xlsx')
        sheet = wb.active

        # Iterate over the rows to find the row with the matching Telegram username
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
            cell = row[0]  # The first column contains Telegram usernames
            if cell.value == telegram_username:
                # If found, update the 4th and 5th columns (D and E) in the same row
                sheet.cell(row=cell.row, column=4).value = username  # 4th column for username
                sheet.cell(row=cell.row, column=5).value = password  # 5th column for password
                break

        # Save the workbook
        wb.save('Book1.xlsx')

        # For debugging purposes

        # Clear the inputs after processing
        context.user_data['inputs'].clear()

        page = ChromiumPage()

        try:
            # Open the target website
            page.get('https://agents.ichancy.com/')

            time.sleep(1)
            page.get('https://agents.ichancy.com/players/players')
            time.sleep(1)
            try:
                username_field = page.ele(
                    'xpath://*[@id="root"]/div/div[3]/div[2]/div[2]/div[2]/form/div[1]/div/label/div[1]/input',
                    timeout=5
                )
                password_field = page.ele(
                    'xpath://*[@id="root"]/div/div[3]/div[2]/div[2]/div[2]/form/div[2]/div/label/div[1]/input',
                    timeout=5
                )

                # If username and password fields are located, fill them
                if username_field and password_field:
                    username_field.input('thelegend@agent.nsp')
                    password_field.input('Aa990@990\n')  # This will submit the form after entering the password
                    sleep(5)
            except Exception:
                # If username/password field is not found, skip filling them
                print("Username or password field not found, skipping login step.")
                pass

            # Add player button
            element = page.ele('css:.btn.playersActionButton-bc')
            element.click()

            time.sleep(1)

            # Fill in player information
            name = page.ele(
                'xpath:.//*[@id="root"]/div/div[4]/div/div/div[2]/form/div/div/div[1]/div/div/label/div[1]/input')
            name.input('ahmed')

            mid_name = page.ele(
                'xpath:.//*[@id="root"]/div/div[4]/div/div/div[2]/form/div/div/div[2]/div/div/label/div[1]/input')
            mid_name.input('ali')

            last_name = page.ele(
                'xpath:.//*[@id="root"]/div/div[4]/div/div/div[2]/form/div/div/div[3]/div/div/label/div[1]/input')
            last_name.input('fateh')

            user_name = page.ele(
                'xpath:.//*[@id="root"]/div/div[4]/div/div/div[2]/form/div/div/div[4]/div/div/label/div[1]/input')
            user_name.input(username)

            phone = page.ele(
                'xpath:.//*[@id="root"]/div/div[4]/div/div/div[2]/form/div/div/div[5]/div/div/label/div[1]/input')
            phone.input('123456789')

            mail = page.ele(
                'xpath:.//*[@id="root"]/div/div[4]/div/div/div[2]/form/div/div/div[6]/div/div/label/div[1]/input')
            mail.input(email)

            keyword = page.ele(
                'xpath:.//*[@id="root"]/div/div[4]/div/div/div[2]/form/div/div/div[7]/div/div/label/div[1]/input')
            keyword.input(password)
            time.sleep(5)
            # Handle the dropdown selection
            dropdown_list = page.ele('xpath:.//input[@placeholder="Countries"]')
            dropdown_list.click()
            andorra_element = page.ele('xpath:.//p[@title="Andorra"]')

            andorra_element.click()

            parent_list = page.ele(
                'xpath:.//*[@id="root"]/div/div[4]/div/div/div[2]/form/div/div/div[10]/div/div[1]/div[1]/label/div[1]/input')
            parent_list.click()

            element = page.ele(
                'xpath:.//*[@id="root"]/div/div[4]/div/div/div[2]/form/div/div/div[10]/div/div[2]/div/button')
            element.click()

            # Submit registration
            reg = page.ele(
                'xpath:.//*[@id="root"]/div/div[4]/div/div/div[3]/button[2]')
            reg.click()

            time.sleep(10)

            await update.message.reply_text(f"بيانات الحساب الخاص بك\n"
                  f"اسم المستخدم: {username}\n"
                  f"كلمة السر: {password}")

        finally:
            # Close the browser
            page.quit()

        # End the conversation
        return ConversationHandler.END



#deposit
def extract_info_from_image(image_path):
    # Load the image
    image = Image.open(image_path)

    # Convert the image to grayscale for better OCR results
    gray_image = image.convert('L')

    # Enhance the image by applying a sharpening filter to make the text clearer
    enhanced_image = gray_image.filter(ImageFilter.SHARPEN)

    # Define the region to crop the first box (for the number on the left)
    first_box_region = enhanced_image.crop((0, 350, image.width // 2, 450))
    # Perform OCR on the entire image to extract all text

    # Extract all occurrences of the text after "Amount:"
    # Define the region to crop the second row (where "To:" is located)
    to_number_region = enhanced_image.crop((0, 450, image.width, 550))

    # Perform OCR on the cropped regions
    ocr_first_box = pytesseract.image_to_string(first_box_region)
    ocr_to_number = pytesseract.image_to_string(to_number_region)
    ocr_result = pytesseract.image_to_string(enhanced_image)


    # Process the OCR result to extract the desired information
    lines = ocr_first_box.splitlines()
    number = lines[0].strip() if lines else "Number not found"
    amounts = re.findall(r'Amount:\s*(\d+)', ocr_result)


    match = re.search(r'To:\s*(\d+)', ocr_to_number)
    to_number = match.group(1) if match else "Number not found"

    return number, to_number , amounts
async def deposit_withdraw(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()

    if query.data == 'deposit':
        keyboard = [
            [InlineKeyboardButton("syriatel cash", callback_data='deposit_cash')],
            [InlineKeyboardButton("payeer", callback_data='deposit_payeer')],
            [InlineKeyboardButton("Bemo bank", callback_data='deposit_bemo')],
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text("اختر طرق الدفع", reply_markup=reply_markup)
async def handle_deposit_method_selection(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()

    if query.data == 'deposit_cash':
        context.user_data['deposit_method'] = query.data.split('_')[1]

        await query.edit_message_text("قم بالتحويل إلى هذا الرقم 22443355 وادخل رقم عملية التحويل")

        context.user_data['awaiting_payment_number'] = True
    elif query.data == 'deposit_payeer':
        await query.edit_message_text("قم بالتحويل إلى هذه المحفظة P1034210265 وادخل المبلغ المودع")
        context.user_data['awaiting_payeerpayment_number'] = True
    elif query.data == 'deposit_bemo':
        await query.edit_message_text("قم بالتحويل إلى هذه المحفظة 22445566 وادخل المبلغ المودع")
        context.user_data['awaiting_bemopayment_number'] = True
#cash handle
processed_payment_numbers = []

# Cash handle function for processing payment numbers
async def message_handler2(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.message.from_user.id

    if context.user_data.get('awaiting_payment_number'):
        payment_number = update.message.text

        # Check if the payment number has already been used
        if payment_number in processed_payment_numbers:
            await update.message.reply_text("تم الايداع مسبقا باستخدام رقم التحويل المدخل")
        else:
            # Store the payment number for this user
            user_payment_number[user_id] = payment_number
            await update.message.reply_text("ادخل رقم المبلغ المراد ايداعه")
            context.user_data['awaiting_payment_number'] = False
            context.user_data['awaiting_deposit_amount'] = True

    elif context.user_data.get('awaiting_deposit_amount'):
        deposit_amount = int(update.message.text)
        if deposit_amount < 15000:
            await update.message.reply_text("المبلغ أقل من الحد الأدنى المسموح به")
        else:
            context.user_data['deposit_amount'] = deposit_amount
            await update.message.reply_text("ارسل لقطة شاشة لعملية التحويل من سيريتل كاش")
            context.user_data['awaiting_deposit_amount'] = False
            context.user_data['awaiting_screenshot'] = True


async def photo_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.message.from_user.id

    if context.user_data.get('awaiting_screenshot'):
        # Check if the message contains a photo
        if not update.message.photo:
            await update.message.reply_text("No photo found in the message. Please send a valid screenshot.")
            return

        # Download the photo
        photo_file = await update.message.photo[-1].get_file()

        # Ensure the 'downloads' directory exists
        downloads_dir = 'downloads'
        if not os.path.exists(downloads_dir):
            os.makedirs(downloads_dir)

        file_path = os.path.join(downloads_dir, f'{user_id}_screenshot.jpg')
        await photo_file.download_to_drive(file_path)

        # Process the image to extract information
        number, to_number, amounts = extract_info_from_image(file_path)

        # Extracting the values from the dictionaries
        deposit_amount = user_deposit_amount[list(user_deposit_amount.keys())[0]]
        payment_number = user_payment_number[list(user_payment_number.keys())[0]]

        # Checking the conditions
        if amounts[0] == deposit_amount and number == payment_number:
            # Update Excel file with the deposit amount
            excel_file = 'Book1.xlsx'
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            # Variables for updating balances
            user_found = False
            referrer_username = None

            # Iterate over rows to find the username and referrer
            for row_index in range(1, ws.max_row + 1):  # Start from row 2
                cell_username = ws.cell(row=row_index, column=1).value  # Username column
                referral_code = ws.cell(row=row_index, column=3).value  # Referral code column

                if cell_username == update.message.from_user.username:
                    # Update the balance for the depositing user
                    ws.cell(row=row_index, column=2, value=amounts[0])  # Balance column
                    user_found = True

                    # Find the referrer
                    if referral_code:
                        for ref_row_index in range(1, ws.max_row + 1):
                            if ws.cell(row=ref_row_index, column=3).value == referral_code:  # Referral code column
                                referrer_username = ws.cell(row=ref_row_index, column=1).value
                                break
                    break

            if user_found and referrer_username:
                # Calculate 10% of the deposit amount
                referral_reward = deposit_amount * 0.10

                # Update the referrer's balance
                for row_index in range(1, ws.max_row + 1):
                    if ws.cell(row=row_index, column=1).value == referrer_username:
                        current_balance = ws.cell(row=row_index, column=2).value or 0
                        ws.cell(row=row_index, column=2, value=current_balance + referral_reward)
                        break

            # Save changes to the Excel file
            wb.save(excel_file)

            # Notify user of successful deposit
            await update.message.reply_text("تم اضافة الرصيد بنجاح")
        else:
            await update.message.reply_text("معلومات الإيداع غير صحيحة. يرجى التحقق من التفاصيل وإعادة الإرسال.")

        # Reset the flag
        context.user_data['awaiting_screenshot'] = False

deposit_amount2 = 0

#payeer handle
async def payeer_handle(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    global deposit_amount2
    deposit_amount2 = int(update.message.text)
    if deposit_amount2 < 15000:
        await update.message.reply_text("المبلغ أقل من الحد الأدنى المسموح به")
    else:
        context.user_data['awaiting_payeerpayment_number'] = deposit_amount2
        await update.message.reply_text("ارسل لقطة شاشة لعملية التحويل من بايير")
        context.user_data['awaiting_payeerpayment_number'] = False
        context.user_data['awaiting_photo'] = True
async def payeer_photo_handle(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    username = update.message.from_user.username

    if context.user_data.get('awaiting_photo'):
        print('fhfhfh')
        # Check if the message contains a photo
        if not update.message.photo:
            await update.message.reply_text("No photo found in the message. Please send a valid screenshot.")
            return

        # Download the photo
        photo_file = await update.message.photo[-1].get_file()

        # Ensure the 'downloads' directory exists
        downloads_dir = 'downloads'
        if not os.path.exists(downloads_dir):
            os.makedirs(downloads_dir)
        file_path = os.path.join(downloads_dir, f'{username}_screenshot.jpg')
        await photo_file.download_to_drive(file_path)
        message_text = f"Amount: {deposit_amount2} SYP\nUsername: @{username}\nPayment method: Payeer"
        recipient_user_id = '5666304947'
        with open(file_path, 'rb') as photo:
            await context.bot.send_photo(chat_id=recipient_user_id, photo=photo, caption=message_text)
        await update.message.reply_text("تم اضافة طلب الايداع سيتم معالجة الطلب")
#bemo bank
deposit_amount3 = 0
async def bemo_handle(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    global deposit_amount3
    deposit_amount3 = int(update.message.text)
    if deposit_amount3 < 15000:
        await update.message.reply_text("المبلغ أقل من الحد الأدنى المسموح به")
    else:
        context.user_data['awaiting_bemopayment_number'] = deposit_amount3
        await update.message.reply_text("ارسل لقطة شاشة لعملية التحويل من تطبيق بنك بيمو")
        context.user_data['awaiting_bemopayment_number'] = False
        context.user_data['awaiting_bemophoto'] = True
async def bemo_photo_handle(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    username = update.message.from_user.username

    if context.user_data.get('awaiting_bemophoto'):
        # Check if the message contains a photo
        if not update.message.photo:
            await update.message.reply_text("No photo found in the message. Please send a valid screenshot.")
            return

        # Download the photo
        photo_file = await update.message.photo[-1].get_file()

        # Ensure the 'downloads' directory exists
        downloads_dir = 'downloads'
        if not os.path.exists(downloads_dir):
            os.makedirs(downloads_dir)
        file_path = os.path.join(downloads_dir, f'{username}_screenshot.jpg')
        await photo_file.download_to_drive(file_path)
        message_text = f"Amount: {deposit_amount3} SYP\nUsername: @{username}\nPayment method: bemo bank"
        recipient_user_id = '5666304947'
        with open(file_path, 'rb') as photo:
            await context.bot.send_photo(chat_id=recipient_user_id, photo=photo, caption=message_text)
        await update.message.reply_text("تم اضافة طلب الايداع سيتم معالجة الطلب")






#wthdraw
# withdraw
async def withdraw(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()

    if query.data == 'withdraw':
        # Show withdrawal options
        keyboard = [
            [InlineKeyboardButton("syriatel cash", callback_data='withdraw_syriatel_cash')],
            [InlineKeyboardButton("payeer", callback_data='withdraw_payeer')],
            [InlineKeyboardButton("Bemo bank", callback_data='withdraw_bemo_bank')],
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text("من فضلك اختر طريقة السحب", reply_markup=reply_markup)
async def handle_withdraw_method_selection(update: Update, context: CallbackContext) -> int:
    query = update.callback_query
    await query.answer()

    if query.data.startswith('withdraw_'):
        context.user_data['withdraw_method'] = query.data.split('_')[1]
        await query.edit_message_text("ادخل رقم التحويل الخاص بك")
        context.user_data['awaiting_wallet_number'] = True
async def collect_wallet_number(update: Update, context: CallbackContext) -> int:
    user_wallet_number = update.message.text
    username = update.effective_user.username

    if context.user_data.get('awaiting_wallet_number'):
        context.user_data['wallet_number'] = update.message.text
        await update.message.reply_text("الان قم بإدخال المبلغ المراد سحبه")
        context.user_data['awaiting_wallet_number'] = False
        context.user_data['awaiting_withdraw_amount'] = True

    # Capture and handle the withdrawal amount
    elif context.user_data.get('awaiting_withdraw_amount'):
        context.user_data['withdraw_amount'] = update.message.text
        withdraw_method = context.user_data.get('withdraw_method')
        wallet_number = context.user_data.get('wallet_number')
        withdraw_amount = float(context.user_data.get('withdraw_amount'))
        excel_data = pd.read_excel("Book1.xlsx")
        user_row = excel_data[excel_data['username'] == username]
        balance = user_row['balance'].values[0]
        if withdraw_amount < 100000:
            await update.message.reply_text("لا يمكن سحب اي مبلغ تحت 100 الف ليرة سورية")
        elif withdraw_amount > balance:
            await update.message.reply_text("هذرا رصيدك الحالي غير كافي لاتمام عملية السحب")
        else:
            new_balance = balance - withdraw_amount
            # Update the balance in the DataFrame
            excel_data.loc[excel_data['username'] == username, 'balance'] = new_balance
            # Write the updated DataFrame back to the Excel file
            excel_data.to_excel("Book1.xlsx", index=False)
            withdraw_after = withdraw_amount * 0.1
            w2 = withdraw_amount - withdraw_after
            await update.message.reply_text(f"جاري معالجة طلب السحب سيصلك مبلغ{w2} ")
            await update.message.reply_text(f'رصيدك الحالي هو {new_balance}')
            await update.message.reply_text(
                f"طريقة السحب: {withdraw_method}\n"
                f"رقم المحفظة: {wallet_number}\n"
                f"المبلغ المراد سحبه: {w2}\n"
                f"اسم المستخدم: @{username}"
            )

            # Reset the state after processing the withdrawal

        context.user_data['awaiting_withdraw_amount'] = False



#my balance
async def my_balance(update: Update, context: CallbackContext) -> None:
    query = update.callback_query
    await query.answer()  # Acknowledge the callback

    user_id = update.effective_user.username

    # Load the Excel file
    df = pd.read_excel('Book1.xlsx')

    # Check if the column name is correct (replace 'User ID' with the correct column name)
    user_row = df[df['username'] == user_id]  # Use the correct column name

    if not user_row.empty:
        balance = user_row.iloc[0]['balance']  # Ensure 'Balance' is also the correct column name
        await query.edit_message_text(f"رصيدك الحالي هو:{balance}")
    else:
        await query.edit_message_text("User not found in the record.")



# reward
async def reward(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()
    await query.edit_message_text("اهلا بك في قسم الجوائز ادخل كود الجائزة")
    context.user_data['awaiting_reward_code'] = True
valid_codes = [
    '482913', '651472', '398205', '720154', '893721',
    '125038', '670492', '385716', '142803', '579128',
    '408392', '923716', '307159', '649287', '593028',
    '184930', '572839', '614023', '853094', '392710'
]

# Function to handle the 'reward' query
async def reward_query_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()
    await query.message.reply_text("اهلا بك في قسم الجوائز ادخل كود الجائزة")
    context.user_data['awaiting_reward_code'] = True
# Function to handle the user's input (reward code)
async def handle_reward_code(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if context.user_data.get('awaiting_reward_code'):
        reward_code = update.message.text.strip()

        # Debugging: Print the entered reward code and the list of valid codes
        print(f"Entered code: {reward_code}")
        print(f"Valid codes: {valid_codes}")

        # Check if the code is valid
        if reward_code in valid_codes:
            # Load the Excel file
            df = pd.read_excel('Book1.xlsx')
            user = update.message.from_user.username

            # Check if the user exists in the Excel file
            if user in df['username'].values:
                index = df[df['username'] == user].index[0]

                # Update the balance
                df.at[index, 'balance'] += 15000

                # Save the updated Excel file
                df.to_excel('Book1.xlsx', index=False)

                # Remove the used code from the list
                valid_codes.remove(reward_code)

                # Send confirmation message
                await update.message.reply_text("مبروك لقد تم تحديث رصيدك")
            else:
                await update.message.reply_text("عذرا، لم يتم العثور على المستخدم في قاعدة البيانات")
        else:
            await update.message.reply_text("الكود غير صحيح")

        # Reset the state
        context.user_data['awaiting_reward_code'] = False





from DrissionPage import ChromiumPage
from time import sleep
def process_deposit(current_username: str, input_amount: float):
    # Initialize DrissionPage
    page = ChromiumPage()
    page.get('https://agents.ichancy.com/')
    sleep(5)

    try:
        username_field = page.ele(
            'xpath://*[@id="root"]/div/div[3]/div[2]/div[2]/div[2]/form/div[1]/div/label/div[1]/input',
            timeout=5
        )
        password_field = page.ele(
            'xpath://*[@id="root"]/div/div[3]/div[2]/div[2]/div[2]/form/div[2]/div/label/div[1]/input',
            timeout=5
        )

        # If username and password fields are located, fill them
        if username_field and password_field:
            username_field.input('thelegend@agent.nsp')
            password_field.input('Aa990@990\n')  # This will submit the form after entering the password
            sleep(5)
    except Exception:
        # If username/password field is not found, skip filling them
        print("Username or password field not found, skipping login step.")
        pass

    # Navigate to the transfer page
    transfer_button = page.ele('css:.headerNavIcon-bc.cursor-pointer .bc-icon-transfer-bold')
    transfer_button.click()
    sleep(5)

    # Locate the user search bar
    user_search_bar = page.ele(
        'xpath://*[@id="root"]/div/div[4]/div/div/div[2]/form/div[1]/div/div[1]/div/div[1]/div[1]/label/div[1]/input'
    )
    sleep(5)

    # Search for the user by username
    user_search_bar.input(current_username)
    sleep(15)

    # Select the correct user from the search results
    buttons = page.eles(
        'css:button.btn.listItem.a-minimal.s-default.f-full-width.c-default.id-start.cr-round'
    )
    for button in buttons:
        if current_username in button.text:
            button.click()
            break

    # Enter the transfer amount
    transfer_amount = page.ele(
        'xpath://*[@id="root"]/div/div[4]/div/div/div[2]/form/div[1]/div/div[2]/div/div/label/div[1]/input'
    )
    transfer_amount.input(str(input_amount))
    sleep(5)

    # Confirm the transfer
    done_button = page.ele('xpath://*[@id="root"]/div/div[4]/div/div/div[3]/button[2]')
    done_button.click()
    sleep(10)

    # Save screenshot
    sleep(3)

    # Close the browser
    page.close()

async def ich_deposit(update: Update, context: CallbackContext) -> int:
    # Ensure we have a callback query
    if update.callback_query:
        # Extract chat_id and user information from callback_query
        chat_id = update.callback_query.message.chat_id
        user = update.callback_query.from_user
        username = user.username

        # Notify user that processing has started
        await context.bot.send_message(chat_id, "We are processing your request. Please wait...")

        # Define the path to your Excel file
        excel_file_path = 'Book1.xlsx'

        # Load the Excel file into a DataFrame
        import pandas as pd
        df = pd.read_excel(excel_file_path)

        # Check if username exists in the first column
        if username in df.iloc[:, 0].values:
            # Locate the row with the username
            row = df[df.iloc[:, 0] == username]

            # Get values from column 4 and column 2
            name = row.iloc[0, 3]  # Column 4
            amount = row.iloc[0, 1]  # Column 2
            row_index = df[df.iloc[:, 0] == username].index[0]
            df.at[row_index, df.columns[1]] = 0  # Column 2
            df.to_excel(excel_file_path, index=False)
            print('ff')

            # Call the process_deposit function with the extracted values
            process_deposit(name, amount)

            # Notify user that the process is complete
            await context.bot.send_message(chat_id, "Your order is done. Thank you!")
        else:
            # Notify user if the username is not found
            await context.bot.send_message(chat_id, f"Username {username} not found in the Excel file.")

    else:
        # Handle the case where there is no callback query
        await context.bot.send_message( "This update does not contain a callback query.")

    return int



EXCEL_FILE = 'Book1.xlsx'
# Function to update the Excel file with the new balance
def update_balance_in_excel(username, amount):
    # Load the workbook and select the active sheet
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active

    # Find the row where the username matches, and update the balance
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
        current_username = row[0].value
        balance_cell = row[1]  # Assuming the balance is in the second column

        if current_username == username:
            current_balance = balance_cell.value
            new_balance = current_balance + int(amount)
            balance_cell.value = new_balance
            wb.save(EXCEL_FILE)  # Save changes to the file
            return True

    return False  # Return False if the username was not found

# Asynchronous function to handle balance updates
async def handle_balance(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if context.user_data.get('awaiting_balance', False):
        message = update.message.text
        pattern = r"Amount:\s*(\d+)\s*SYP\s*Username:\s*@(\w+)\s*Payment method:\s*(.*)"
        match = re.search(pattern, message)

        if match:
            amount = match.group(1)
            username = match.group(2)
            # Update the balance in the Excel file
            success = update_balance_in_excel(username, amount)

            if success:
                await update.message.reply_text("تمت اضافة الرصيد بنجاح")
            else:
                await update.message.reply_text(f"Username @{username} not found in the system.")

        else:
            await update.message.reply_text("Please provide the message in the correct format.")

        # Reset the flag after processing the message
        context.user_data['awaiting_balance'] = False
    else:
        # If the bot is not expecting a message, it ignores or handles messages normally
        await update.message.reply_text("This message does not follow the expected format.")

# Asynchronous function to handle the /addbalance command
async def add_balance(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    context.user_data['awaiting_balance'] = True
    await update.message.reply_text("اضف نص رسالة طلب الايداع فقط سيتم اضافة الرصيد بشكل اوتوماتيكي")

# Function to handle user messages after /addbalance is issued




# Function to handle the user's input (reward code)
def main() -> None:
    application = Application.builder().token("7016588209:AAHlKI3foDHlT07OV_XDg6XdWDjtX7XenWg").build()
    # Set up the conversation handler
    conv_handler = ConversationHandler(
        entry_points=[CallbackQueryHandler(create_account, pattern='create_account')],
        states={
            COLLECTING_INPUTS: [MessageHandler(filters.TEXT & ~filters.COMMAND, collect_inputs)],
        },
        fallbacks=[],
        per_message=False,
    )

    conv_handler2 = ConversationHandler(
        entry_points=[CommandHandler('start', start)],
        states={
            ASK_REFERRAL: [CallbackQueryHandler(ask_referral)],
            GET_REFERRAL: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_referral)],
            NO_REFERRAL: [MessageHandler(filters.TEXT & ~filters.COMMAND, no_referral)],
        },
        fallbacks=[CommandHandler('start', start)],
    )



    # Add handlers

    application.add_handler(CallbackQueryHandler(deposit_withdraw, pattern='^deposit$'))
    application.add_handler(CallbackQueryHandler(handle_deposit_method_selection, pattern='^deposit_'))
    application.add_handler(CallbackQueryHandler(withdraw, pattern='^withdraw$'))
    application.add_handler(CallbackQueryHandler(handle_withdraw_method_selection, pattern='^withdraw_'))
    application.add_handler(CallbackQueryHandler(my_balance, pattern='^balance$'))
    application.add_handler(CallbackQueryHandler(reward, pattern='^reward'))
    application.add_handler(CallbackQueryHandler(ich_deposit, pattern='^ich_deposit'))
    application.add_handler(CommandHandler("addbalance", add_balance))


    # Add both message handlers in a single line with 'elif' for collecting inputs
    application.add_handler(
        MessageHandler(
            filters.TEXT & ~filters.COMMAND,
            lambda update, context: (
                message_handler2(update, context) if 'awaiting_payment_number' in context.user_data else
                collect_wallet_number(update, context) if 'awaiting_wallet_number' in context.user_data else
                payeer_handle(update, context) if 'awaiting_payeerpayment_number' in context.user_data else
                bemo_handle(update, context) if 'awaiting_bemopayment_number' in context.user_data else
                handle_reward_code(update, context) if 'awaiting_reward_code' in context.user_data else
                collect_inputs(update, context) if 'inputs' in context.user_data else
                get_referral(update, context) if 'referral_code' in context.user_data else
                handle_balance(update, context) if 'awaiting_balance' in context.user_data else
                no_referral(update, context)
            )
        )
    )

    application.add_handler(
        MessageHandler(
            filters.PHOTO & ~filters.COMMAND,
            lambda update, context: (
                photo_handler(update, context) if 'awaiting_screenshot' in context.user_data
                else bemo_photo_handle(update, context) if 'awaiting_bemophoto' in context.user_data
                else payeer_photo_handle(update, context)

            )
        )
    )






    application.add_handler(conv_handler)
    application.add_handler(conv_handler2)

    # Start polling
    application.run_polling()

if __name__ == '__main__':
    main()






from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler, MessageHandler, ConversationHandler,
    ContextTypes, filters
)
from PIL import Image, ImageFilter
import pytesseract
import re
import os
import pandas as pd
import openpyxl
import random
import string
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackContext
from selenium import webdriver
from seleniumbase import Driver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from selenium.webdriver.common.keys import Keys

COLLECTING_INPUTS = range(1)
COLLECTING_DATA = range(1)


# Define  variables
user_payment_number = {}
user_deposit_amount = {}
# generate referral number
def generate_referral_code(length=8):
    letters_and_digits = string.ascii_letters + string.digits
    return ''.join(random.choice(letters_and_digits) for _ in range(length))




# State definitions for ConversationHandler
ASK_REFERRAL, GET_REFERRAL,NO_REFERRAL = range(3)


# Function to start the bot and ask about referral code
async def start(update: Update, context: CallbackContext) -> int:
    df = pd.read_excel('Book1.xlsx')
    user = update.message.from_user
    username = user.username
    if username in df['username'].values:
        return await show(update, context)
    else:
        referral_message = "هل تملك رقم احالة من احد اضدقاؤك؟"

        # Create the referral keyboard (Yes / No)
        referral_keyboard = [
            [InlineKeyboardButton("نعم", callback_data='yes_referral')],
            [InlineKeyboardButton("لا", callback_data='no_referral')]
        ]
        reply_markup = InlineKeyboardMarkup(referral_keyboard)

        # Ask user if they have a referral code
        await update.message.reply_text(referral_message, reply_markup=reply_markup)


    return ASK_REFERRAL



# Callback to handle the user's referral answer
async def ask_referral(update: Update, context: CallbackContext) -> int:
    query = update.callback_query
    await query.answer()

    if query.data == 'yes_referral':
        await query.edit_message_text("الرجاء إدخال رقم الاحالة:")
        return GET_REFERRAL  # Move to the next state to get the referral code

    elif query.data == 'no_referral':
        # Proceed without referral code, show main menu
        await no_referral(update, context)
        return NO_REFERRAL


# Function to get the referral code

async def get_referral(update: Update, context: CallbackContext) -> int:
    referral_code = update.message.text
    context.user_data['referral_code'] = referral_code  # Save referral code

    # Path to the Excel file
    excel_file = 'Book1.xlsx'

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    referral_column = 3
    username_column = 1  # Column index for usernames
    new_row = []  # List to store new row data

    # Check if the referral code exists in the column
    referral_exists = False
    for row in ws.iter_rows(min_col=referral_column, max_col=referral_column, values_only=True):
        if row[0] == referral_code:
            referral_exists = True
            break

    if referral_exists:
        await update.message.reply_text(f"تم حفظ رقم الاحالة: {referral_code}")

        # Generate a new referral code (for demonstration purposes, use UUID)
        new_referral_code = generate_referral_code()
        # Retrieve the current username of the user
        username = update.message.from_user.username

        # Add new user details to a new row in Excel
        new_row = [username, "", new_referral_code]  # Assuming the columns are: username, some column, referral code, new referral code

        ws.append(new_row)  # Append the new row to the worksheet
        wb.save(excel_file)  # Save the changes to the Excel file

        # Proceed to show the main menu
        await show(update, context)

    else:
        await update.message.reply_text("لقد ادخلت رمز دعوى خاطئ")

    return ConversationHandler.END



async def no_referral(update: Update, context: CallbackContext) -> None:
    excel_file = 'Book1.xlsx'
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    new_referral_code = generate_referral_code()
    # Retrieve the current username of the user
    if update.message and update.message.from_user:
        username = update.message.from_user.username
    elif update.callback_query and update.callback_query.from_user:
        username = update.callback_query.from_user.username
    else:
        username = "Unknown"

    # Add new user details to a new row in Excel
    new_row = [username, "",
               new_referral_code]  # Assuming the columns are: username, some column, referral code, new referral code

    ws.append(new_row)  # Append the new row to the worksheet
    wb.save(excel_file)  # Save the changes to the Excel file

    # Proceed to show the main menu
    await show(update, context)



# Function to show the main menu
async def show(update: Update, context: CallbackContext) -> None:
    welcome_message = (
        "أهلا بك في بوت ايتشانسي\n"
        "يمكنك هذا البوت من إنشاء حساب أو تعبئة رصيد والاطلاع على سجل الرهانات الخاصة بك "
    )

    keyboard = [
        [InlineKeyboardButton("🎓انشاء حساب ايتشانسي", callback_data='create_account')],
        [InlineKeyboardButton(" تعبئة وسحب اينشانسي", callback_data='deposit')],
        [InlineKeyboardButton("سحب اينشانسي", callback_data='withdraw')],
        [InlineKeyboardButton("رصيدي", callback_data='balance')],
        [InlineKeyboardButton("كود جائزة", callback_data='reward')],
        [InlineKeyboardButton("جروب الدعم", url="https://t.me/+JkZ3-g6U7oM0NGQ0")],
        [InlineKeyboardButton("📊 اايداع الرصيد في ايتشانسي", callback_data='ich_deposit')],
        [InlineKeyboardButton("⚙️ السحل", callback_data='history')],
        [InlineKeyboardButton("💎 الدليل الشامل", callback_data='subscription')],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    # Check if the update comes from a message or a callback query
    if update.message:
        await update.message.reply_text(welcome_message, reply_markup=reply_markup)
    elif update.callback_query:
        await update.callback_query.message.edit_text(welcome_message, reply_markup=reply_markup)


async def create_account(update: Update, context: CallbackContext) -> int:
    query = update.callback_query
    await query.answer()
    await query.edit_message_text("ادخل اسم المستخدم، كلمة المرور،الايميل, كل منهما في رسالة منفصلة:")
    await collect_inputs(update, context)


    return COLLECTING_INPUTS
# Collect inputs function
async def collect_inputs(update: Update, context: CallbackContext) -> int:
    # Initialize the inputs list if it doesn't exist yet
    if 'inputs' not in context.user_data:
        context.user_data['inputs'] = []

    # Append the user input to the list
    context.user_data['inputs'].append(update.message.text)

    # Once we have collected all three inputs (username, password, email)
    if len(context.user_data['inputs']) == 3:
        username, password, email = context.user_data['inputs']

        # Get the Telegram username of the user
        telegram_username = update.message.from_user.username

        # Reply to the user
        await update.message.reply_text("حاري انشاء الحساب الخاص بك يرجى الانتظار!")

        # Update the Excel file with the collected inputs
        wb = openpyxl.load_workbook('Book1.xlsx')
        sheet = wb.active

        # Iterate over the rows to find the row with the matching Telegram username
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
            cell = row[0]  # The first column contains Telegram usernames
            if cell.value == telegram_username:
                # If found, update the 4th and 5th columns (D and E) in the same row
                sheet.cell(row=cell.row, column=4).value = username  # 4th column for username
                sheet.cell(row=cell.row, column=5).value = password  # 5th column for password
                break

        # Save the workbook
        wb.save('Book1.xlsx')

        # For debugging purposes

        # Clear the inputs after processing
        context.user_data['inputs'].clear()

        page = ChromiumPage()

        try:
            # Open the target website
            page.get('https://agents.ichancy.com/')

            time.sleep(1)
            page.get('https://agents.ichancy.com/players/players')
            time.sleep(1)
            try:
                username_field = page.ele(
                    'xpath://*[@id="root"]/div/div[3]/div[2]/div[2]/div[2]/form/div[1]/div/label/div[1]/input',
                    timeout=5
                )
                password_field = page.ele(
                    'xpath://*[@id="root"]/div/div[3]/div[2]/div[2]/div[2]/form/div[2]/div/label/div[1]/input',
                    timeout=5
                )

                # If username and password fields are located, fill them
                if username_field and password_field:
                    username_field.input('thelegend@agent.nsp')
                    password_field.input('Aa990@990\n')  # This will submit the form after entering the password
                    sleep(5)
            except Exception:
                # If username/password field is not found, skip filling them
                print("Username or password field not found, skipping login step.")
                pass

            # Add player button
            element = page.ele('css:.btn.playersActionButton-bc')
            element.click()

            time.sleep(1)

            # Fill in player information
            name = page.ele(
                'xpath:.//*[@id="root"]/div/div[4]/div/div/div[2]/form/div/div/div[1]/div/div/label/div[1]/input')
            name.input('ahmed')

            mid_name = page.ele(
                'xpath:.//*[@id="root"]/div/div[4]/div/div/div[2]/form/div/div/div[2]/div/div/label/div[1]/input')
            mid_name.input('ali')

            last_name = page.ele(
                'xpath:.//*[@id="root"]/div/div[4]/div/div/div[2]/form/div/div/div[3]/div/div/label/div[1]/input')
            last_name.input('fateh')

            user_name = page.ele(
                'xpath:.//*[@id="root"]/div/div[4]/div/div/div[2]/form/div/div/div[4]/div/div/label/div[1]/input')
            user_name.input(username)

            phone = page.ele(
                'xpath:.//*[@id="root"]/div/div[4]/div/div/div[2]/form/div/div/div[5]/div/div/label/div[1]/input')
            phone.input('123456789')

            mail = page.ele(
                'xpath:.//*[@id="root"]/div/div[4]/div/div/div[2]/form/div/div/div[6]/div/div/label/div[1]/input')
            mail.input(email)

            keyword = page.ele(
                'xpath:.//*[@id="root"]/div/div[4]/div/div/div[2]/form/div/div/div[7]/div/div/label/div[1]/input')
            keyword.input(password)
            time.sleep(5)
            # Handle the dropdown selection
            dropdown_list = page.ele('xpath:.//input[@placeholder="Countries"]')
            dropdown_list.click()
            andorra_element = page.ele('xpath:.//p[@title="Andorra"]')

            andorra_element.click()

            parent_list = page.ele(
                'xpath:.//*[@id="root"]/div/div[4]/div/div/div[2]/form/div/div/div[10]/div/div[1]/div[1]/label/div[1]/input')
            parent_list.click()

            element = page.ele(
                'xpath:.//*[@id="root"]/div/div[4]/div/div/div[2]/form/div/div/div[10]/div/div[2]/div/button')
            element.click()

            # Submit registration
            reg = page.ele(
                'xpath:.//*[@id="root"]/div/div[4]/div/div/div[3]/button[2]')
            reg.click()

            time.sleep(10)

            await update.message.reply_text(f"بيانات الحساب الخاص بك\n"
                  f"اسم المستخدم: {username}\n"
                  f"كلمة السر: {password}")

        finally:
            # Close the browser
            page.quit()

        # End the conversation
        return ConversationHandler.END



#deposit
def extract_info_from_image(image_path):
    # Load the image
    image = Image.open(image_path)

    # Convert the image to grayscale for better OCR results
    gray_image = image.convert('L')

    # Enhance the image by applying a sharpening filter to make the text clearer
    enhanced_image = gray_image.filter(ImageFilter.SHARPEN)

    # Define the region to crop the first box (for the number on the left)
    first_box_region = enhanced_image.crop((0, 350, image.width // 2, 450))
    # Perform OCR on the entire image to extract all text

    # Extract all occurrences of the text after "Amount:"
    # Define the region to crop the second row (where "To:" is located)
    to_number_region = enhanced_image.crop((0, 450, image.width, 550))

    # Perform OCR on the cropped regions
    ocr_first_box = pytesseract.image_to_string(first_box_region)
    ocr_to_number = pytesseract.image_to_string(to_number_region)
    ocr_result = pytesseract.image_to_string(enhanced_image)


    # Process the OCR result to extract the desired information
    lines = ocr_first_box.splitlines()
    number = lines[0].strip() if lines else "Number not found"
    amounts = re.findall(r'Amount:\s*(\d+)', ocr_result)


    match = re.search(r'To:\s*(\d+)', ocr_to_number)
    to_number = match.group(1) if match else "Number not found"

    return number, to_number , amounts
async def deposit_withdraw(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()

    if query.data == 'deposit':
        keyboard = [
            [InlineKeyboardButton("syriatel cash", callback_data='deposit_cash')],
            [InlineKeyboardButton("payeer", callback_data='deposit_payeer')],
            [InlineKeyboardButton("Bemo bank", callback_data='deposit_bemo')],
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text("اختر طرق الدفع", reply_markup=reply_markup)
async def handle_deposit_method_selection(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()

    if query.data == 'deposit_cash':
        context.user_data['deposit_method'] = query.data.split('_')[1]

        await query.edit_message_text("قم بالتحويل إلى هذا الرقم 22443355 وادخل رقم عملية التحويل")

        context.user_data['awaiting_payment_number'] = True
    elif query.data == 'deposit_payeer':
        await query.edit_message_text("قم بالتحويل إلى هذه المحفظة P1034210265 وادخل المبلغ المودع")
        context.user_data['awaiting_payeerpayment_number'] = True
    elif query.data == 'deposit_bemo':
        await query.edit_message_text("قم بالتحويل إلى هذه المحفظة 22445566 وادخل المبلغ المودع")
        context.user_data['awaiting_bemopayment_number'] = True
#cash handle
processed_payment_numbers = []

# Cash handle function for processing payment numbers
async def message_handler2(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.message.from_user.id

    if context.user_data.get('awaiting_payment_number'):
        payment_number = update.message.text

        # Check if the payment number has already been used
        if payment_number in processed_payment_numbers:
            await update.message.reply_text("تم الايداع مسبقا باستخدام رقم التحويل المدخل")
        else:
            # Store the payment number for this user
            user_payment_number[user_id] = payment_number
            await update.message.reply_text("ادخل رقم المبلغ المراد ايداعه")
            context.user_data['awaiting_payment_number'] = False
            context.user_data['awaiting_deposit_amount'] = True

    elif context.user_data.get('awaiting_deposit_amount'):
        deposit_amount = int(update.message.text)
        if deposit_amount < 15000:
            await update.message.reply_text("المبلغ أقل من الحد الأدنى المسموح به")
        else:
            context.user_data['deposit_amount'] = deposit_amount
            await update.message.reply_text("ارسل لقطة شاشة لعملية التحويل من سيريتل كاش")
            context.user_data['awaiting_deposit_amount'] = False
            context.user_data['awaiting_screenshot'] = True


async def photo_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.message.from_user.id

    if context.user_data.get('awaiting_screenshot'):
        # Check if the message contains a photo
        if not update.message.photo:
            await update.message.reply_text("No photo found in the message. Please send a valid screenshot.")
            return

        # Download the photo
        photo_file = await update.message.photo[-1].get_file()

        # Ensure the 'downloads' directory exists
        downloads_dir = 'downloads'
        if not os.path.exists(downloads_dir):
            os.makedirs(downloads_dir)

        file_path = os.path.join(downloads_dir, f'{user_id}_screenshot.jpg')
        await photo_file.download_to_drive(file_path)

        # Process the image to extract information
        number, to_number, amounts = extract_info_from_image(file_path)

        # Extracting the values from the dictionaries
        deposit_amount = user_deposit_amount[list(user_deposit_amount.keys())[0]]
        payment_number = user_payment_number[list(user_payment_number.keys())[0]]

        # Checking the conditions
        if amounts[0] == deposit_amount and number == payment_number:
            # Update Excel file with the deposit amount
            excel_file = 'Book1.xlsx'
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            # Variables for updating balances
            user_found = False
            referrer_username = None

            # Iterate over rows to find the username and referrer
            for row_index in range(1, ws.max_row + 1):  # Start from row 2
                cell_username = ws.cell(row=row_index, column=1).value  # Username column
                referral_code = ws.cell(row=row_index, column=3).value  # Referral code column

                if cell_username == update.message.from_user.username:
                    # Update the balance for the depositing user
                    ws.cell(row=row_index, column=2, value=amounts[0])  # Balance column
                    user_found = True

                    # Find the referrer
                    if referral_code:
                        for ref_row_index in range(1, ws.max_row + 1):
                            if ws.cell(row=ref_row_index, column=3).value == referral_code:  # Referral code column
                                referrer_username = ws.cell(row=ref_row_index, column=1).value
                                break
                    break

            if user_found and referrer_username:
                # Calculate 10% of the deposit amount
                referral_reward = deposit_amount * 0.10

                # Update the referrer's balance
                for row_index in range(1, ws.max_row + 1):
                    if ws.cell(row=row_index, column=1).value == referrer_username:
                        current_balance = ws.cell(row=row_index, column=2).value or 0
                        ws.cell(row=row_index, column=2, value=current_balance + referral_reward)
                        break

            # Save changes to the Excel file
            wb.save(excel_file)

            # Notify user of successful deposit
            await update.message.reply_text("تم اضافة الرصيد بنجاح")
        else:
            await update.message.reply_text("معلومات الإيداع غير صحيحة. يرجى التحقق من التفاصيل وإعادة الإرسال.")

        # Reset the flag
        context.user_data['awaiting_screenshot'] = False

deposit_amount2 = 0

#payeer handle
async def payeer_handle(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    global deposit_amount2
    deposit_amount2 = int(update.message.text)
    if deposit_amount2 < 15000:
        await update.message.reply_text("المبلغ أقل من الحد الأدنى المسموح به")
    else:
        context.user_data['awaiting_payeerpayment_number'] = deposit_amount2
        await update.message.reply_text("ارسل لقطة شاشة لعملية التحويل من بايير")
        context.user_data['awaiting_payeerpayment_number'] = False
        context.user_data['awaiting_photo'] = True
async def payeer_photo_handle(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    username = update.message.from_user.username

    if context.user_data.get('awaiting_photo'):
        print('fhfhfh')
        # Check if the message contains a photo
        if not update.message.photo:
            await update.message.reply_text("No photo found in the message. Please send a valid screenshot.")
            return

        # Download the photo
        photo_file = await update.message.photo[-1].get_file()

        # Ensure the 'downloads' directory exists
        downloads_dir = 'downloads'
        if not os.path.exists(downloads_dir):
            os.makedirs(downloads_dir)
        file_path = os.path.join(downloads_dir, f'{username}_screenshot.jpg')
        await photo_file.download_to_drive(file_path)
        message_text = f"Amount: {deposit_amount2} SYP\nUsername: @{username}\nPayment method: Payeer"
        recipient_user_id = '5666304947'
        with open(file_path, 'rb') as photo:
            await context.bot.send_photo(chat_id=recipient_user_id, photo=photo, caption=message_text)
        await update.message.reply_text("تم اضافة طلب الايداع سيتم معالجة الطلب")
#bemo bank
deposit_amount3 = 0
async def bemo_handle(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    global deposit_amount3
    deposit_amount3 = int(update.message.text)
    if deposit_amount3 < 15000:
        await update.message.reply_text("المبلغ أقل من الحد الأدنى المسموح به")
    else:
        context.user_data['awaiting_bemopayment_number'] = deposit_amount3
        await update.message.reply_text("ارسل لقطة شاشة لعملية التحويل من تطبيق بنك بيمو")
        context.user_data['awaiting_bemopayment_number'] = False
        context.user_data['awaiting_bemophoto'] = True
async def bemo_photo_handle(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    username = update.message.from_user.username

    if context.user_data.get('awaiting_bemophoto'):
        # Check if the message contains a photo
        if not update.message.photo:
            await update.message.reply_text("No photo found in the message. Please send a valid screenshot.")
            return

        # Download the photo
        photo_file = await update.message.photo[-1].get_file()

        # Ensure the 'downloads' directory exists
        downloads_dir = 'downloads'
        if not os.path.exists(downloads_dir):
            os.makedirs(downloads_dir)
        file_path = os.path.join(downloads_dir, f'{username}_screenshot.jpg')
        await photo_file.download_to_drive(file_path)
        message_text = f"Amount: {deposit_amount3} SYP\nUsername: @{username}\nPayment method: bemo bank"
        recipient_user_id = '5666304947'
        with open(file_path, 'rb') as photo:
            await context.bot.send_photo(chat_id=recipient_user_id, photo=photo, caption=message_text)
        await update.message.reply_text("تم اضافة طلب الايداع سيتم معالجة الطلب")






#wthdraw
# withdraw
async def withdraw(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()

    if query.data == 'withdraw':
        # Show withdrawal options
        keyboard = [
            [InlineKeyboardButton("syriatel cash", callback_data='withdraw_syriatel_cash')],
            [InlineKeyboardButton("payeer", callback_data='withdraw_payeer')],
            [InlineKeyboardButton("Bemo bank", callback_data='withdraw_bemo_bank')],
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text("من فضلك اختر طريقة السحب", reply_markup=reply_markup)
async def handle_withdraw_method_selection(update: Update, context: CallbackContext) -> int:
    query = update.callback_query
    await query.answer()

    if query.data.startswith('withdraw_'):
        context.user_data['withdraw_method'] = query.data.split('_')[1]
        await query.edit_message_text("ادخل رقم التحويل الخاص بك")
        context.user_data['awaiting_wallet_number'] = True
async def collect_wallet_number(update: Update, context: CallbackContext) -> int:
    user_wallet_number = update.message.text
    username = update.effective_user.username

    if context.user_data.get('awaiting_wallet_number'):
        context.user_data['wallet_number'] = update.message.text
        await update.message.reply_text("الان قم بإدخال المبلغ المراد سحبه")
        context.user_data['awaiting_wallet_number'] = False
        context.user_data['awaiting_withdraw_amount'] = True

    # Capture and handle the withdrawal amount
    elif context.user_data.get('awaiting_withdraw_amount'):
        context.user_data['withdraw_amount'] = update.message.text
        withdraw_method = context.user_data.get('withdraw_method')
        wallet_number = context.user_data.get('wallet_number')
        withdraw_amount = float(context.user_data.get('withdraw_amount'))
        excel_data = pd.read_excel("Book1.xlsx")
        user_row = excel_data[excel_data['username'] == username]
        balance = user_row['balance'].values[0]
        if withdraw_amount < 100000:
            await update.message.reply_text("لا يمكن سحب اي مبلغ تحت 100 الف ليرة سورية")
        elif withdraw_amount > balance:
            await update.message.reply_text("هذرا رصيدك الحالي غير كافي لاتمام عملية السحب")
        else:
            new_balance = balance - withdraw_amount
            # Update the balance in the DataFrame
            excel_data.loc[excel_data['username'] == username, 'balance'] = new_balance
            # Write the updated DataFrame back to the Excel file
            excel_data.to_excel("Book1.xlsx", index=False)
            withdraw_after = withdraw_amount * 0.1
            w2 = withdraw_amount - withdraw_after
            await update.message.reply_text(f"جاري معالجة طلب السحب سيصلك مبلغ{w2} ")
            await update.message.reply_text(f'رصيدك الحالي هو {new_balance}')
            await update.message.reply_text(
                f"طريقة السحب: {withdraw_method}\n"
                f"رقم المحفظة: {wallet_number}\n"
                f"المبلغ المراد سحبه: {w2}\n"
                f"اسم المستخدم: @{username}"
            )

            # Reset the state after processing the withdrawal

        context.user_data['awaiting_withdraw_amount'] = False



#my balance
async def my_balance(update: Update, context: CallbackContext) -> None:
    query = update.callback_query
    await query.answer()  # Acknowledge the callback

    user_id = update.effective_user.username

    # Load the Excel file
    df = pd.read_excel('Book1.xlsx')

    # Check if the column name is correct (replace 'User ID' with the correct column name)
    user_row = df[df['username'] == user_id]  # Use the correct column name

    if not user_row.empty:
        balance = user_row.iloc[0]['balance']  # Ensure 'Balance' is also the correct column name
        await query.edit_message_text(f"رصيدك الحالي هو:{balance}")
    else:
        await query.edit_message_text("User not found in the record.")



# reward
async def reward(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()
    await query.edit_message_text("اهلا بك في قسم الجوائز ادخل كود الجائزة")
    context.user_data['awaiting_reward_code'] = True
valid_codes = [
    '482913', '651472', '398205', '720154', '893721',
    '125038', '670492', '385716', '142803', '579128',
    '408392', '923716', '307159', '649287', '593028',
    '184930', '572839', '614023', '853094', '392710'
]

# Function to handle the 'reward' query
async def reward_query_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()
    await query.message.reply_text("اهلا بك في قسم الجوائز ادخل كود الجائزة")
    context.user_data['awaiting_reward_code'] = True
# Function to handle the user's input (reward code)
async def handle_reward_code(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if context.user_data.get('awaiting_reward_code'):
        reward_code = update.message.text.strip()

        # Debugging: Print the entered reward code and the list of valid codes
        print(f"Entered code: {reward_code}")
        print(f"Valid codes: {valid_codes}")

        # Check if the code is valid
        if reward_code in valid_codes:
            # Load the Excel file
            df = pd.read_excel('Book1.xlsx')
            user = update.message.from_user.username

            # Check if the user exists in the Excel file
            if user in df['username'].values:
                index = df[df['username'] == user].index[0]

                # Update the balance
                df.at[index, 'balance'] += 15000

                # Save the updated Excel file
                df.to_excel('Book1.xlsx', index=False)

                # Remove the used code from the list
                valid_codes.remove(reward_code)

                # Send confirmation message
                await update.message.reply_text("مبروك لقد تم تحديث رصيدك")
            else:
                await update.message.reply_text("عذرا، لم يتم العثور على المستخدم في قاعدة البيانات")
        else:
            await update.message.reply_text("الكود غير صحيح")

        # Reset the state
        context.user_data['awaiting_reward_code'] = False





from DrissionPage import ChromiumPage
from time import sleep
def process_deposit(current_username: str, input_amount: float):
    # Initialize DrissionPage
    page = ChromiumPage()
    page.get('https://agents.ichancy.com/')
    sleep(5)

    try:
        username_field = page.ele(
            'xpath://*[@id="root"]/div/div[3]/div[2]/div[2]/div[2]/form/div[1]/div/label/div[1]/input',
            timeout=5
        )
        password_field = page.ele(
            'xpath://*[@id="root"]/div/div[3]/div[2]/div[2]/div[2]/form/div[2]/div/label/div[1]/input',
            timeout=5
        )

        # If username and password fields are located, fill them
        if username_field and password_field:
            username_field.input('thelegend@agent.nsp')
            password_field.input('Aa990@990\n')  # This will submit the form after entering the password
            sleep(5)
    except Exception:
        # If username/password field is not found, skip filling them
        print("Username or password field not found, skipping login step.")
        pass

    # Navigate to the transfer page
    transfer_button = page.ele('css:.headerNavIcon-bc.cursor-pointer .bc-icon-transfer-bold')
    transfer_button.click()
    sleep(5)

    # Locate the user search bar
    user_search_bar = page.ele(
        'xpath://*[@id="root"]/div/div[4]/div/div/div[2]/form/div[1]/div/div[1]/div/div[1]/div[1]/label/div[1]/input'
    )
    sleep(5)

    # Search for the user by username
    user_search_bar.input(current_username)
    sleep(15)

    # Select the correct user from the search results
    buttons = page.eles(
        'css:button.btn.listItem.a-minimal.s-default.f-full-width.c-default.id-start.cr-round'
    )
    for button in buttons:
        if current_username in button.text:
            button.click()
            break

    # Enter the transfer amount
    transfer_amount = page.ele(
        'xpath://*[@id="root"]/div/div[4]/div/div/div[2]/form/div[1]/div/div[2]/div/div/label/div[1]/input'
    )
    transfer_amount.input(str(input_amount))
    sleep(5)

    # Confirm the transfer
    done_button = page.ele('xpath://*[@id="root"]/div/div[4]/div/div/div[3]/button[2]')
    done_button.click()
    sleep(10)

    # Save screenshot
    sleep(3)

    # Close the browser
    page.close()

async def ich_deposit(update: Update, context: CallbackContext) -> int:
    # Ensure we have a callback query
    if update.callback_query:
        # Extract chat_id and user information from callback_query
        chat_id = update.callback_query.message.chat_id
        user = update.callback_query.from_user
        username = user.username

        # Notify user that processing has started
        await context.bot.send_message(chat_id, "We are processing your request. Please wait...")

        # Define the path to your Excel file
        excel_file_path = 'Book1.xlsx'

        # Load the Excel file into a DataFrame
        import pandas as pd
        df = pd.read_excel(excel_file_path)

        # Check if username exists in the first column
        if username in df.iloc[:, 0].values:
            # Locate the row with the username
            row = df[df.iloc[:, 0] == username]

            # Get values from column 4 and column 2
            name = row.iloc[0, 3]  # Column 4
            amount = row.iloc[0, 1]  # Column 2
            row_index = df[df.iloc[:, 0] == username].index[0]
            df.at[row_index, df.columns[1]] = 0  # Column 2
            df.to_excel(excel_file_path, index=False)
            print('ff')

            # Call the process_deposit function with the extracted values
            process_deposit(name, amount)

            # Notify user that the process is complete
            await context.bot.send_message(chat_id, "Your order is done. Thank you!")
        else:
            # Notify user if the username is not found
            await context.bot.send_message(chat_id, f"Username {username} not found in the Excel file.")

    else:
        # Handle the case where there is no callback query
        await context.bot.send_message( "This update does not contain a callback query.")

    return int



EXCEL_FILE = 'Book1.xlsx'
# Function to update the Excel file with the new balance
def update_balance_in_excel(username, amount):
    # Load the workbook and select the active sheet
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active

    # Find the row where the username matches, and update the balance
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
        current_username = row[0].value
        balance_cell = row[1]  # Assuming the balance is in the second column

        if current_username == username:
            current_balance = balance_cell.value
            new_balance = current_balance + int(amount)
            balance_cell.value = new_balance
            wb.save(EXCEL_FILE)  # Save changes to the file
            return True

    return False  # Return False if the username was not found

# Asynchronous function to handle balance updates
async def handle_balance(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if context.user_data.get('awaiting_balance', False):
        message = update.message.text
        pattern = r"Amount:\s*(\d+)\s*SYP\s*Username:\s*@(\w+)\s*Payment method:\s*(.*)"
        match = re.search(pattern, message)

        if match:
            amount = match.group(1)
            username = match.group(2)
            # Update the balance in the Excel file
            success = update_balance_in_excel(username, amount)

            if success:
                await update.message.reply_text("تمت اضافة الرصيد بنجاح")
            else:
                await update.message.reply_text(f"Username @{username} not found in the system.")

        else:
            await update.message.reply_text("Please provide the message in the correct format.")

        # Reset the flag after processing the message
        context.user_data['awaiting_balance'] = False
    else:
        # If the bot is not expecting a message, it ignores or handles messages normally
        await update.message.reply_text("This message does not follow the expected format.")

# Asynchronous function to handle the /addbalance command
async def add_balance(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    context.user_data['awaiting_balance'] = True
    await update.message.reply_text("اضف نص رسالة طلب الايداع فقط سيتم اضافة الرصيد بشكل اوتوماتيكي")

# Function to handle user messages after /addbalance is issued




# Function to handle the user's input (reward code)
def main() -> None:
    application = Application.builder().token("7016588209:AAHlKI3foDHlT07OV_XDg6XdWDjtX7XenWg").build()
    # Set up the conversation handler
    conv_handler = ConversationHandler(
        entry_points=[CallbackQueryHandler(create_account, pattern='create_account')],
        states={
            COLLECTING_INPUTS: [MessageHandler(filters.TEXT & ~filters.COMMAND, collect_inputs)],
        },
        fallbacks=[],
        per_message=False,
    )

    conv_handler2 = ConversationHandler(
        entry_points=[CommandHandler('start', start)],
        states={
            ASK_REFERRAL: [CallbackQueryHandler(ask_referral)],
            GET_REFERRAL: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_referral)],
            NO_REFERRAL: [MessageHandler(filters.TEXT & ~filters.COMMAND, no_referral)],
        },
        fallbacks=[CommandHandler('start', start)],
    )



    # Add handlers

    application.add_handler(CallbackQueryHandler(deposit_withdraw, pattern='^deposit$'))
    application.add_handler(CallbackQueryHandler(handle_deposit_method_selection, pattern='^deposit_'))
    application.add_handler(CallbackQueryHandler(withdraw, pattern='^withdraw$'))
    application.add_handler(CallbackQueryHandler(handle_withdraw_method_selection, pattern='^withdraw_'))
    application.add_handler(CallbackQueryHandler(my_balance, pattern='^balance$'))
    application.add_handler(CallbackQueryHandler(reward, pattern='^reward'))
    application.add_handler(CallbackQueryHandler(ich_deposit, pattern='^ich_deposit'))
    application.add_handler(CommandHandler("addbalance", add_balance))


    # Add both message handlers in a single line with 'elif' for collecting inputs
    application.add_handler(
        MessageHandler(
            filters.TEXT & ~filters.COMMAND,
            lambda update, context: (
                message_handler2(update, context) if 'awaiting_payment_number' in context.user_data else
                collect_wallet_number(update, context) if 'awaiting_wallet_number' in context.user_data else
                payeer_handle(update, context) if 'awaiting_payeerpayment_number' in context.user_data else
                bemo_handle(update, context) if 'awaiting_bemopayment_number' in context.user_data else
                handle_reward_code(update, context) if 'awaiting_reward_code' in context.user_data else
                collect_inputs(update, context) if 'inputs' in context.user_data else
                get_referral(update, context) if 'referral_code' in context.user_data else
                handle_balance(update, context) if 'awaiting_balance' in context.user_data else
                no_referral(update, context)
            )
        )
    )

    application.add_handler(
        MessageHandler(
            filters.PHOTO & ~filters.COMMAND,
            lambda update, context: (
                photo_handler(update, context) if 'awaiting_screenshot' in context.user_data
                else bemo_photo_handle(update, context) if 'awaiting_bemophoto' in context.user_data
                else payeer_photo_handle(update, context)

            )
        )
    )






    application.add_handler(conv_handler)
    application.add_handler(conv_handler2)

    # Start polling
    application.run_polling()

if __name__ == '__main__':
    main()

















